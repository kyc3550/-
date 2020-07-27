from tkinter import *
from tkinter.ttk import *
from datetime import datetime
import tkinter.messagebox
import openpyxl
import os

class MainPage():
    today= datetime.today().strftime("%Y-%m-%d")
    day = datetime.today().day
    month = datetime.today().month
    success = "님이 출석되었습니다"
    members_path=os.path.abspath('출석현황.xlsx')
    
    members = openpyxl.load_workbook("출석현황.xlsx")
    members_sheet = members['%s월'% month]
        
    def checkin(self):
        namelist=[]
        name_all = self.members_sheet['B']
        select_name=name.get()

        for cell in name_all:
            namelist.append(cell.value)
            if cell.value==select_name:
                cnt=len(namelist)
                self.members_sheet.cell(cnt,self.day+2,"O")
                tkinter.messagebox.showinfo("성공",name.get()+self.success)
        self.members.save(r'%s'%self.members_path)
        name.delete(0,END)

    def reseach(self):
        frame = Frame(app)
        frame.grid(row=2,column=1)
        
        treeview=tkinter.ttk.Treeview(frame, columns=["one", "two"], displaycolumns=[ "one","two"],height=31)
        treeview.pack(side="left")

        vsb = tkinter.Scrollbar(frame, orient="vertical", command=treeview.yview)
        vsb.pack(side='right', fill='y')

        treeview.configure(yscrollcommand=vsb.set)

        treeview.column("#0", width=100)
        treeview.heading("#0", text="번호")
        
        treeview.column("#1", width=100)
        treeview.heading("two", text="출결")

        treeview.column("#2", width=100)
        treeview.heading("one", text="이름")

        i=3
        treelist=[]
        while i<len(self.members_sheet['B']):
            treelist.append([(self.members_sheet.cell(i,2).value),(self.members_sheet.cell(i,self.day+2).value)])
            i=i+1
        for data in range(len(treelist)):
            treeview.insert('','end',text=data,values=treelist[data])
        

    def menu_add(self):
        new=AddNewUser()
    
    def __init__(self):
        global app
        app=Tk()
        app.title("출석 체크")

        menubar = tkinter.Menu(app)   
        helpmenu = tkinter.Menu(menubar)
        helpmenu.add_command(label="추가",command=self.menu_add)  
        menubar.add_cascade(label="추가", menu=helpmenu)  
        app.config(menu=menubar) 
        
        info=["이름"]
        a=1
        Label(app,text=self.today,width=10,font=(25)).grid(row=0,column=1)
        for c in info:
            Label(app,text=c,width=10,font=(25)).grid(row=a,column=0)
            a=a+1
            
        global name
        name=Entry(app,width=20,font=(25))
        name.grid(row=1,column=1)

        Button(app,text="출석",width='10',command=self.checkin).grid(row=1,column=2)
        Button(app,text="새로고침",width='10',command=self.reseach).grid(row=0,column=2)

class AddNewUser(MainPage):
    member_info=openpyxl.load_workbook("회원연락처.xlsx")
    member_info_sheet=member_info['회원연락처']
    path=os.path.abspath('회원연락처.xlsx')
    def save(self):
        self.member_info_sheet.append([new_name.get(),phonenum.get(),address.get(),period.get()])
        self.member_info.save(r'%s'%self.path)
        tkinter.messagebox.showinfo("완료",new_name.get()+"님 등록 완료")
        new_name.delete(0,END)
        phonenum.delete(0,END)
        address.delete(0,END)
        period.delete(0,END)
    def update(self):
        name_all = self.member_info_sheet['A']
        data=[]
        for cell in name_all:
            data.append(cell.value)

        i=1
        row=3
        for item in data[1:]:
            MainPage.members_sheet['B'+str(row)]=data[i]
            i=i+1
            row=row+1
        MainPage.members.save(r'%s'%MainPage.members_path)
        tkinter.messagebox.showinfo("완료","새로고침완료")
        
        
    def __init__(self):
        app2=Tk()
        app2.title("회원가입")

        info=["이름","연락처","주소","등록기간"]
        Label(app2,text="신규회원 등록",width=15,font=(25)).grid(row=0,column=1)
        
        a=1
        for c in info:
            Label(app2,text=c,width=10,font=(25)).grid(row=a,column=0)
            a=a+1
        
        global new_name
        new_name=Entry(app2,width=15,font=(25))
        new_name.grid(row=1,column=1)

        global phonenum
        phonenum=Entry(app2,width=15,font=(25))
        phonenum.grid(row=2,column=1)

        global address
        address=Entry(app2,width=15,font=(25))
        address.grid(row=3,column=1)

        global period
        period=Entry(app2,width=25,font=(25))
        period.grid(row=4,column=1,padx=5)
        
        Button(app2,text="등록",width='10',command=self.save).grid(row=1,column=2)
        Button(app2,text="추가",width='10', command=self.update).grid(row=2,column=2,padx=5)


play=MainPage()
