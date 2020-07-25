from tkinter import *
from tkinter.ttk import *
from datetime import datetime
import tkinter.messagebox
import openpyxl
import os

today= datetime.today().strftime("%Y-%m-%d")

class Taepung():   
    info=["이름"]
    a=1
    day = datetime.today().day
    month = datetime.today().month
    success = "님이 출석되었습니다"
    
    def find(self):
        path=os.path.abspath('출석현황.xlsx')
        namelist = "출석현황.xlsx"
        members = openpyxl.load_workbook(namelist)
        members_sheet = members['%s월'%self.month]
        namelist=[]
        name_all = members_sheet['B']
        select_name=name.get()
        for cell in name_all:
            namelist.append(cell.value)
            if cell.value==select_name:
                cnt=len(namelist)
                members_sheet.cell(cnt,self.day+2,"O")
                tkinter.messagebox.showinfo("성공",name.get()+self.success)
        members.save(r'%s'%path)
        name.delete(0,END)
        
        
    def helpbar(self):
        new=NewUser()
        
    def reseach(self):
        frame = Frame(app)
        frame.grid(row=2,column=1)
        
        treeview=tkinter.ttk.Treeview(frame, columns=["one", "two"], displaycolumns=[ "one","two"],height=31)
        treeview.pack(side="left")

        vsb = tkinter.Scrollbar(frame, orient="vertical", command=treeview.yview)
        vsb.pack(side='right', fill='y')

        treeview.configure(yscrollcommand=vsb.set)

        treeview.column("#0", width=100)
        treeview.heading("#0", text="번호")
        
        treeview.column("#1", width=100)
        treeview.heading("two", text="출결")

        treeview.column("#2", width=100)
        treeview.heading("one", text="이름")
        
        namelist = "출석현황.xlsx"
        
        members = openpyxl.load_workbook(namelist)
        members_sheet = members['%s월'%self.month]
        
        i=3
        treelist=[]
        while i<200:
            #(members_sheet.cell(i,1).value)
            treelist.append([(members_sheet.cell(i,2).value),(members_sheet.cell(i,self.day+2).value)])
            i=i+1
        for a in range(len(treelist)):
            treeview.insert('','end',text=a,values=treelist[a])
        members.close()

        
    def __init__(self):
        global app
        app=Tk()
        app.title("출석 체크")

        menubar = tkinter.Menu(app)   
        helpmenu = tkinter.Menu(menubar)
        helpmenu.add_command(label="추가",command=self.helpbar)  
        menubar.add_cascade(label="추가", menu=helpmenu)  
        app.config(menu=menubar) 
        
        
        Label(app,text=today,width=10,font=(25)).grid(row=0,column=1)
        for c in self.info:
            Label(app,text=c,width=10,font=(25)).grid(row=self.a,column=0)
            self.a=self.a+1
            
        global name
        name=Entry(app,width=20,font=(25))
        name.grid(row=1,column=1)

        Button(app,text="출석",width='10',command=self.find).grid(row=1,column=2)
        Button(app,text="새로고침",width='10',command=self.reseach).grid(row=0,column=2)

        
        
        
        
        app.mainloop()




class NewUser():
    info=["이름","연락처","주소"]
    a=1
    success = "저장완료"
    month = datetime.today().month
    def save(self):
        path=os.path.abspath('회원연락처.xlsx')
        add=openpyxl.load_workbook("회원연락처.xlsx")
        add_sheet=add['회원연락처']
        add_sheet.append([new_name.get(),phonenum.get(),address.get()])
        add.save(r'%s'%path)
        tkinter.messagebox.showinfo("완료",new_name.get()+self.success)
        new_name.delete(0,END)
        phonenum.delete(0,END)
        address.delete(0,END)
        
    def update(self):
        path=os.path.abspath('출석현황.xlsx')
        add=openpyxl.load_workbook("회원연락처.xlsx")
        add_sheet=add['회원연락처']
        name_all = add_sheet['A']
        
        global data
        data=[]
        for cell in name_all:
            data.append(cell.value)
        attendnce = openpyxl.load_workbook("출석현황.xlsx")
        attendnce_sheet = attendnce['%s월'%self.month]
        i=1
        row=3
        for item in data[1:]:
            attendnce_sheet['B'+str(row)]=data[i]
            i=i+1
            row=row+1
        attendnce.save(r'%s'%path)
        tkinter.messagebox.showinfo("완료","새로고침완료")
        
    def __init__(self):
        app2=Tk()
        app2.title("회원가입")
        app2.geometry('400x130+1000+600')
        
        Label(app2,text="신규회원 등록",width=15,font=(25)).grid(row=0,column=1)
        for c in self.info:
            Label(app2,text=c,width=10,font=(25)).grid(row=self.a,column=0)
            self.a=self.a+1
    
        global new_name
        new_name=Entry(app2,width=15,font=(25))
        new_name.grid(row=1,column=1)

        global phonenum
        phonenum=Entry(app2,width=15,font=(25))
        phonenum.grid(row=2,column=1)

        global address
        address=Entry(app2,width=15,font=(25))
        address.grid(row=3,column=1)

        Button(app2,text="저장",width='10',command=self.save).grid(row=1,column=2)
        Button(app2,text="추가",width='10', command=self.update).grid(row=2,column=2,padx=5)

        
        
play=Taepung()
